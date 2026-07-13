from pathlib import Path

from openpyxl import load_workbook

from aip_canonica.internal.models import (
    Workbook,
    Sheet,
    Row,
    Cell,
    CellLocation,
)

from .document_parser import DocumentParser


class ExcelParser(DocumentParser):

    def parse(
        self,
        path: Path,
    ) -> Workbook:
        """Parses an Excel file into a Workbook model.

        Args:
            path (Path): The file path to the Excel workbook.

        Returns:
            Workbook: The parsed workbook model.
        """
        workbook = Workbook()

        excel_workbook = load_workbook(
            filename=path,
            data_only=False,
        )

        for worksheet in excel_workbook.worksheets:

            workbook.sheets.append(self._parse_sheet(worksheet))

        return workbook

    def _parse_sheet(self, worksheet) -> Sheet:
        """Parses an Excel worksheet into a Sheet model.

        Args:
            worksheet: The openpyxl worksheet object.

        Returns:
            Sheet: The parsed sheet model.
        """
        sheet = Sheet(
            name=worksheet.title,
        )

        for cells in worksheet.iter_rows():
            sheet.rows.append(self._parse_row(worksheet.title, cells))

        return sheet

    def _parse_row(self, sheet_name: str, cells: tuple) -> Row:
        """Parses a row of cells into a Row model.

        Args:
            sheet_name (str): The name of the sheet.
            cells (tuple): A tuple of openpyxl cell objects representing the row.

        Returns:
            Row: The parsed row model.
        """
        return Row(
            index=cells[0].row,
            cells=[self._parse_cell(sheet_name, cell) for cell in cells],
        )

    def _parse_cell(self, sheet_name: str, cell: 'openpyxl.cell.cell.Cell') -> Cell:
        """Parses a cell into a Cell model.

        Args:
            sheet_name (str): The name of the sheet.
            cell (openpyxl.cell.cell.Cell): The openpyxl cell object.

        Returns:
            Cell: The parsed cell model.
        """
        return Cell(
            value=cell.value,
            location=CellLocation(
                sheet=sheet_name,
                row=cell.row,
                column=cell.column,
                address=cell.coordinate,
            ),
        )